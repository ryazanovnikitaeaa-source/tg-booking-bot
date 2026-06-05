import asyncio
import csv
import logging
import os
import uuid
from datetime import datetime, timedelta, date

from dotenv import load_dotenv
load_dotenv()

from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command, CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    Message,
    ReplyKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardRemove,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    CallbackQuery,
)
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import (
    MASTER_NAME,
    MASTER_USERNAME,
    WORK_START,
    WORK_END,
    MAX_PER_DAY,
    MIN_INTERVAL,
    DAYS_AHEAD,
    WELCOME_TEXT,
)

# ─── Конфигурация ────────────────────────────────────────────────────────────
BOT_TOKEN = os.getenv("BOT_TOKEN")
OWNER_ID  = int(os.getenv("OWNER_ID"))
XLSX_FILE    = "zayavki.xlsx"
OLD_CSV_FILE = "zayavki.csv"

# Внутренние ключи и русские заголовки для xlsx
FIELDS = ["id", "created_at", "service", "name", "date", "time",
          "telegram_id", "username"]
HEADERS_RU = ["ID", "Дата записи", "Услуга", "Имя", "Дата",
               "Время", "Telegram ID", "Username"]
COL_WIDTHS = [10, 18, 14, 20, 13, 8, 14, 20]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

# ─── FSM ─────────────────────────────────────────────────────────────────────
class BookingStates(StatesGroup):
    waiting_for_name = State()
    waiting_for_date = State()
    waiting_for_time = State()

# ─── Локализация ─────────────────────────────────────────────────────────────
MONTHS_RU   = ["января","февраля","марта","апреля","мая","июня",
                "июля","августа","сентября","октября","ноября","декабря"]
WEEKDAYS_RU = ["Пн","Вт","Ср","Чт","Пт","Сб","Вс"]

# ─── Клавиатуры ──────────────────────────────────────────────────────────────
SERVICES_KB = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📝 Записаться")],
        [KeyboardButton(text="💬 Связаться с мастером")],
        [KeyboardButton(text="❌ Отменить запись")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Выберите действие...",
)

CANCEL_KB = ReplyKeyboardMarkup(
    keyboard=[[KeyboardButton(text="❌ Отмена")]],
    resize_keyboard=True,
)

def get_calendar_kb() -> InlineKeyboardMarkup:
    today = date.today()
    rows, row = [], []
    for i in range(DAYS_AHEAD):
        d = today + timedelta(days=i)
        label = f"{d.day:02d}.{d.month:02d} {WEEKDAYS_RU[d.weekday()]}"
        row.append(InlineKeyboardButton(
            text=label, callback_data=f"date:{d.strftime('%Y-%m-%d')}"
        ))
        if len(row) == 3:
            rows.append(row); row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton(text="❌ Отмена", callback_data="cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)

def get_time_slots_kb(date_str: str) -> InlineKeyboardMarkup | None:
    """Генерирует инлайн-кнопки доступных слотов времени для выбранной даты.
    Слоты: 09:00–16:00, шаг 30 мин. Исключает прошедшее время и занятые слоты
    (с учётом минимального интервала MIN_INTERVAL между записями).
    Возвращает None, если нет ни одного свободного слота.
    """
    now = datetime.now()

    # Собираем занятые минуты по уже существующим записям на эту дату
    booked_minutes = []
    for b in get_day_bookings(date_str):
        try:
            t = datetime.strptime(b["time"], "%H:%M").time()
            booked_minutes.append(t.hour * 60 + t.minute)
        except (ValueError, KeyError):
            pass

    available = []
    for hour in range(WORK_START, WORK_END + 1):
        for minute in (0, 30):
            if hour == WORK_END and minute > 0:
                break  # 16:30 и позже — не показываем
            slot = f"{hour:02d}:{minute:02d}"
            slot_dt = datetime.strptime(f"{date_str} {slot}", "%Y-%m-%d %H:%M")
            if slot_dt <= now:
                continue  # прошедшее время
            slot_min = hour * 60 + minute
            if any(abs(slot_min - bm) < MIN_INTERVAL for bm in booked_minutes):
                continue  # конфликт по интервалу
            available.append(slot)

    if not available:
        return None

    # Строим клавиатуру — 3 кнопки в ряд
    rows, row = [], []
    for slot in available:
        row.append(InlineKeyboardButton(text=slot, callback_data=f"timeslot:{slot}"))
        if len(row) == 3:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton(text="❌ Отмена", callback_data="cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def fmt_date(date_str: str) -> str:
    d = datetime.strptime(date_str, "%Y-%m-%d").date()
    return f"{d.day} {MONTHS_RU[d.month - 1]}"

# ─── XLSX: чтение / запись ───────────────────────────────────────────────────

def _style_header(ws):
    """Красит шапку таблицы."""
    blue  = PatternFill("solid", fgColor="2E75B6")
    white = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = blue
        cell.font = white
        cell.alignment = center
    ws.row_dimensions[1].height = 20

def read_all_bookings() -> list:
    if not os.path.exists(XLSX_FILE):
        return []
    wb = load_workbook(XLSX_FILE, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return []
    # Первая строка — заголовки, дальше — данные
    return [
        {FIELDS[i]: (str(row[i]) if row[i] is not None else "")
         for i in range(len(FIELDS))}
        for row in rows[1:]
        if any(v is not None for v in row)
    ]

def write_all_bookings(bookings: list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"
    ws.append(HEADERS_RU)
    for b in bookings:
        ws.append([b.get(f, "") for f in FIELDS])
    _style_header(ws)
    for idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width
    try:
        wb.save(XLSX_FILE)
    except PermissionError:
        logger.error("Не могу сохранить %s — файл открыт в другой программе!", XLSX_FILE)

def save_booking(booking_id, service, name, date_str, time_str, telegram_id, username):
    bookings = read_all_bookings()
    bookings.append({
        "id":           booking_id,
        "created_at":  datetime.now().strftime("%Y-%m-%d %H:%M"),
        "service":     service,
        "name":        name,
        "date":        date_str,
        "time":        time_str,
        "telegram_id": str(telegram_id),
        "username":    username or "",
    })
    write_all_bookings(bookings)

# ─── XLSX: миграция из CSV ───────────────────────────────────────────────────
MONTHS_RU_MAP = {m: i+1 for i, m in enumerate(MONTHS_RU)}

def _parse_ru_date(date_display: str) -> str:
    """'10 июня' → '2025-06-10'  (текущий год)"""
    parts = date_display.strip().split()
    if len(parts) == 2:
        day, month_name = parts
        month_num = MONTHS_RU_MAP.get(month_name.lower())
        if month_num:
            year = datetime.now().year
            return f"{year}-{month_num:02d}-{int(day):02d}"
    return date_display   # оставляем как есть, если не распарсилось

def migrate_csv_to_xlsx():
    """Если есть zayavki.csv но нет zayavki.xlsx — конвертируем."""
    if os.path.exists(XLSX_FILE) or not os.path.exists(OLD_CSV_FILE):
        return
    logger.info("Мигрируем %s → %s ...", OLD_CSV_FILE, XLSX_FILE)
    with open(OLD_CSV_FILE, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    bookings = []
    for r in rows:
        # Определяем формат: старый (5 колонок) или новый (8 колонок)
        if "id" in r:
            bookings.append({f: r.get(f, "") for f in FIELDS})
        else:
            date_raw = r.get("Дата", "")
            date_iso = _parse_ru_date(date_raw) if date_raw else ""
            bookings.append({
                "id":           str(uuid.uuid4())[:8],
                "created_at":  r.get("Дата записи", ""),
                "service":     r.get("Услуга", ""),
                "name":        r.get("Имя", ""),
                "date":        date_iso,
                "time":        r.get("Время", ""),
                "telegram_id": "",
                "username":    "",
            })

    write_all_bookings(bookings)
    logger.info("Миграция завершена: %d записей перенесено в xlsx.", len(bookings))

# ─── Бизнес-логика: проверки слотов ─────────────────────────────────────────

def get_day_bookings(date_str: str) -> list:
    return [b for b in read_all_bookings() if b.get("date") == date_str]

def check_slot(date_str: str, time_str: str) -> str | None:
    """
    Возвращает None если слот свободен, иначе строку с причиной отказа.
    Проверяет:
      1. Рабочее время (09:00–16:00)
      2. Лимит 5 записей в день
      3. Интервал ≥ 60 минут с ближайшими записями
    """
    try:
        time_obj = datetime.strptime(time_str, "%H:%M").time()
    except ValueError:
        return "format"

    total_min = time_obj.hour * 60 + time_obj.minute

    if not (WORK_START * 60 <= total_min <= WORK_END * 60):
        return "hours"

    day_bookings = get_day_bookings(date_str)

    if len(day_bookings) >= MAX_PER_DAY:
        return "full"

    for b in day_bookings:
        try:
            existing = datetime.strptime(b["time"], "%H:%M").time()
            diff = abs(total_min - (existing.hour * 60 + existing.minute))
            if diff < MIN_INTERVAL:
                return f"interval:{b['time']}"
        except (ValueError, KeyError):
            pass

    return None

def get_today_bookings() -> list:
    return get_day_bookings(date.today().strftime("%Y-%m-%d"))

def get_user_future_bookings(telegram_id: int) -> list:
    now = datetime.now()
    result = []
    for b in read_all_bookings():
        if b.get("telegram_id") != str(telegram_id):
            continue
        try:
            dt = datetime.strptime(f"{b['date']} {b['time']}", "%Y-%m-%d %H:%M")
            if dt > now:
                result.append(b)
        except (ValueError, KeyError):
            pass
    return result

# ─── Инициализация ───────────────────────────────────────────────────────────
bot = Bot(token=BOT_TOKEN)
dp  = Dispatcher(storage=MemoryStorage())

@dp.startup()
async def on_startup():
    migrate_csv_to_xlsx()
    logger.info("Бот готов к работе.")


# ─── Хэндлеры ────────────────────────────────────────────────────────────────

@dp.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    await state.clear()
    await message.answer(
        WELCOME_TEXT,
        reply_markup=SERVICES_KB,
    )

# /admin — только владелец
@dp.message(Command("admin"))
async def cmd_admin(message: Message):
    if message.from_user.id != OWNER_ID:
        await message.answer("⛔ Нет доступа.")
        return

    now = datetime.now()
    end_date_str = (now.date() + timedelta(days=DAYS_AHEAD - 1)).strftime("%Y-%m-%d")

    valid_bookings = []
    for b in read_all_bookings():
        try:
            dt = datetime.strptime(f"{b['date']} {b['time']}", "%Y-%m-%d %H:%M")
            if dt > now and b["date"] <= end_date_str:
                valid_bookings.append(b)
        except (ValueError, KeyError):
            pass

    if not valid_bookings:
        await message.answer(f"📋 Записей на ближайшие {DAYS_AHEAD} дней нет.")
        return

    # Сортируем
    valid_bookings.sort(key=lambda b: (b.get("date", ""), b.get("time", "")))

    # Группируем по дате
    grouped = {}
    for b in valid_bookings:
        d_str = b["date"]
        if d_str not in grouped:
            grouped[d_str] = []
        grouped[d_str].append(b)

    lines = []
    for d_str in sorted(grouped.keys()):
        if lines:
            lines.append("")
        lines.append(f"📅 {fmt_date(d_str)}")
        for idx, b in enumerate(grouped[d_str], 1):
            un = f" (@{b['username']})" if b.get("username") else ""
            lines.append(f"{idx}. 🕐 {b['time']} — {b['name']}{un}")

    await message.answer("\n".join(lines))

# Отмена записи клиентом (инлайн-кнопка)
@dp.callback_query(F.data.startswith("cancel_"))
async def cancel_booking(callback: CallbackQuery):
    booking_id = callback.data[len("cancel_"):]

    bookings = read_all_bookings()
    booking = next((b for b in bookings if b.get("id") == booking_id), None)

    if not booking:
        await callback.answer("Запись не найдена или уже отменена.", show_alert=True)
        return

    # Проверяем что запись принадлежит этому пользователю
    if booking.get("telegram_id") != str(callback.from_user.id):
        await callback.answer("Это не ваша запись.", show_alert=True)
        return

    # Удаляем из xlsx
    write_all_bookings([b for b in bookings if b.get("id") != booking_id])

    # Подтверждение клиенту
    await callback.message.edit_text("✅ Запись отменена. Ждём вас снова!")
    await callback.answer()

    # Уведомление владельцу
    date_display = fmt_date(booking["date"]) if booking.get("date") else "—"
    un = f"(@{booking['username']})" if booking.get("username") else ""
    try:
        await bot.send_message(
            OWNER_ID,
            f"❌ Отмена записи:\n\n"
            f"👤 {booking['name']} {un}\n"
            f"📅 {date_display}, 🕐 {booking['time']}",
        )
    except Exception as e:
        logger.error("Не удалось уведомить владельца об отмене: %s", e)


# Отмена записи (reply-кнопка — сброс FSM)
@dp.message(F.text == "❌ Отмена")
async def cancel_reply(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("❌ Запись отменена.\n\nВыберите действие:", reply_markup=SERVICES_KB)

# Отмена (инлайн — в календаре)
@dp.callback_query(F.data == "cancel")
async def cancel_inline(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text("❌ Запись отменена.")
    await callback.message.answer("Выберите действие:", reply_markup=SERVICES_KB)
    await callback.answer()


# Связаться с мастером (без FSM, без записи)
@dp.message(F.text == "💬 Связаться с мастером")
async def contact_master(message: Message):
    kb = InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(text="✉️ Написать мастеру", url=f"https://t.me/{MASTER_USERNAME}")
    ]])
    await message.answer(
        f"По всем вопросам напишите мастеру:\n\n👉 @{MASTER_USERNAME}",
        reply_markup=kb,
    )

# Отменить запись (reply-кнопка главного меню)
@dp.message(F.text == "❌ Отменить запись")
async def cancel_booking_menu(message: Message):
    bookings = sorted(
        get_user_future_bookings(message.from_user.id),
        key=lambda b: (b.get("date", ""), b.get("time", ""))
    )
    if not bookings:
        await message.answer("У вас нет активных записей.")
        return
    await message.answer(f"📋 Ваши активные записи ({len(bookings)}):")
    for b in bookings:
        date_display = fmt_date(b["date"]) if b.get("date") else "—"
        username_part = f" (@{b['username']})" if b.get("username") else ""
        text = (
            f"📅 {date_display}, 🕐 {b['time']}\n"
            f"👤 {b['name']}{username_part}"
        )
        kb = InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="❌ Отменить", callback_data=f"cancel_{b['id']}")
        ]])
        await message.answer(text, reply_markup=kb)

# Нажатие на кнопку записи
@dp.message(F.text == "📝 Записаться")
async def service_chosen(message: Message, state: FSMContext):
    # Ограничение: не более 1 активной записи на клиента
    if get_user_future_bookings(message.from_user.id):
        await message.answer(
            "⚠️ У вас уже есть активная запись. Отмените её, прежде чем записаться снова.\n\n"
            "Нажмите *❌ Отменить запись* в меню.",
            parse_mode="Markdown",
        )
        return
    await state.update_data(service=message.text)
    await state.set_state(BookingStates.waiting_for_name)
    await message.answer(
        "Введите ваше имя:",
        reply_markup=CANCEL_KB,
    )

# Ввод имени
@dp.message(BookingStates.waiting_for_name)
async def name_received(message: Message, state: FSMContext):
    name = message.text.strip()
    if not name:
        await message.answer("Имя не может быть пустым. Попробуйте ещё раз:")
        return
    await state.update_data(name=name)
    await state.set_state(BookingStates.waiting_for_date)
    await message.answer(
        f"Приятно познакомиться, *{name}*! 😊\n\nВыберите удобную дату:",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove(),
    )
    await message.answer(f"📅 Ближайшие {DAYS_AHEAD} дней:", reply_markup=get_calendar_kb())

# Выбор даты
@dp.callback_query(F.data.startswith("date:"), BookingStates.waiting_for_date)
async def date_chosen(callback: CallbackQuery, state: FSMContext):
    date_str     = callback.data.split(":")[1]
    date_display = fmt_date(date_str)

    # Проверяем лимит мест уже на этапе выбора даты
    if len(get_day_bookings(date_str)) >= MAX_PER_DAY:
        await callback.answer(
            f"❌ На {date_display} мест нет — все {MAX_PER_DAY} слотов заняты.",
            show_alert=True,
        )
        return

    slots_kb = get_time_slots_kb(date_str)
    if slots_kb is None:
        await callback.answer(
            f"😔 На {date_display} нет свободных слотов. Выберите другую дату.",
            show_alert=True,
        )
        return

    await state.update_data(date=date_str, date_display=date_display)
    await state.set_state(BookingStates.waiting_for_time)
    await callback.message.edit_text(
        f"📅 *{date_display}* — выберите удобное время:",
        parse_mode="Markdown",
        reply_markup=slots_kb,
    )
    await callback.answer()


# Выбор слота времени через инлайн-кнопку
@dp.callback_query(F.data.startswith("timeslot:"), BookingStates.waiting_for_time)
async def timeslot_chosen(callback: CallbackQuery, state: FSMContext):
    time_str = callback.data.split(":", 1)[1]
    data         = await state.get_data()
    date_str     = data["date"]
    date_display = data["date_display"]

    # Защита от гонки: перепроверяем слот в момент нажатия
    appt_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
    if appt_dt <= datetime.now():
        await callback.answer("⚠️ Это время уже прошло!", show_alert=True)
        return
    issue = check_slot(date_str, time_str)
    if issue:
        await callback.answer("⚠️ Этот слот только что заняли. Выберите другое время.", show_alert=True)
        return

    # Всё ок — сохраняем
    service    = data["service"]
    name       = data["name"]
    username   = callback.from_user.username
    booking_id = str(uuid.uuid4())[:8]

    await state.clear()
    save_booking(booking_id, service, name, date_str, time_str, callback.from_user.id, username or "")

    # Подтверждение клиенту
    await callback.message.edit_text(
        "✅ *Запись подтверждена!*\n\n"
        f"👤 Имя: {name}\n"
        f"📅 Дата: {date_display}\n"
        f"🕐 Время: {time_str}\n\n"
        "Мы ждём вас! Если планы изменятся — пожалуйста, сообщите заранее 🙏",
        parse_mode="Markdown",
    )
    await callback.message.answer("Выберите действие:", reply_markup=SERVICES_KB)
    await callback.answer()

    # Уведомление владельцу
    tg_link = f"@{username}" if username else "нет username"
    try:
        await bot.send_message(
            OWNER_ID,
            "🔔 Новая запись!\n\n"
            f"👤 Имя: {name}\n"
            f"📅 Дата: {date_display}\n"
            f"🕐 Время: {time_str}\n"
            f"💬 Telegram: {tg_link} (ID: {callback.from_user.id})\n"
            f"🆔 ID записи: {booking_id}\n\n"
            "📋 Все записи: /admin",
        )
    except Exception as e:
        logger.error("Не удалось уведомить владельца: %s", e)


# ─── Запуск ──────────────────────────────────────────────────────────────────
async def main():
    logger.info("Бот запускается...")
    retry_delay = 5
    while True:
        try:
            await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())
            break  # нормальное завершение (Ctrl+C)
        except Exception as e:
            logger.error("Ошибка подключения: %s", e)
            logger.info("Повтор через %d сек...", retry_delay)
            await asyncio.sleep(retry_delay)
            retry_delay = min(retry_delay * 2, 60)  # экспоненциальный backoff до 60 с

if __name__ == "__main__":
    asyncio.run(main())
